import tkinter as tk
from tkinter import messagebox, simpledialog, colorchooser, font
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
from flask import Flask, render_template, request, redirect, url_for, flash, session
import openpyxl
from openpyxl import Workbook
import os
import webbrowser  # Make sure to import webbrowser at the top
import tkinter.ttk as ttk  # Import ttk for Treeview
import time  # Import time to handle the lockout duration

def login():
    form = LoginForm()
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and bcrypt.check_password_hash(user.password, password):
            session['user_id'] = user.id
            session.permanent = True
            flash('Logged in successfully.', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Login unsuccessful. Please check username and password.', 'danger')
    return render_template('login.html', form=form)

class Account:
    def __init__(self, username, password):
        self.username = username
        self.password = password

class CreateAccountWindow:
    def __init__(self, master, login_window):
        self.master = master
        self.login_window = login_window
        self.master.title("Create Account")
        self.master.geometry("300x150")

        self.username_label = tk.Label(master, text="Username:")
        self.username_label.pack()
        self.username_entry = tk.Entry(master)
        self.username_entry.pack()

        self.password_label = tk.Label(master, text="Password:")
        self.password_label.pack()
        self.password_entry = tk.Entry(master, show="*")
        self.password_entry.pack()

        self.security_code_label = tk.Label(master, text="Security Code:")
        self.security_code_label.pack()
        self.security_code_entry = tk.Entry(master, show="*")
        self.security_code_entry.pack()

        self.create_button = tk.Button(master, text="Create Account", command=self.create_account)
        self.create_button.pack(pady=10)

    def create_account(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        security_code = self.security_code_entry.get()  # New line to get security code

        if username and password and security_code:  # Updated condition to check security code
            if self.login_window.account_exists(username):
                messagebox.showerror("Account Creation Failed", "Username already exists")
            else:
                self.login_window.add_account(username, password, security_code)  # Pass security code
                messagebox.showinfo("Account Created", "Your account has been created successfully")
                self.master.destroy()  # Close the create account window
        else:
            messagebox.showerror("Account Creation Failed", "Please enter username, password, and security code")

class LoginWindow:
    def __init__(self, master):
        self.master = master
        self.master.title("Login")
        self.master.geometry("300x200")

        self.accounts_file = "accounts.xlsx"
        self.ensure_accounts_file_exists()

        self.username_label = tk.Label(master, text="Username:")
        self.username_label.pack()
        self.username_entry = tk.Entry(master)
        self.username_entry.pack()

        self.password_label = tk.Label(master, text="Password:")
        self.password_label.pack()
        self.password_entry = tk.Entry(master, show="*")
        self.password_entry.pack()

        self.login_button = tk.Button(master, text="Login", command=self.login)
        self.login_button.pack(pady=5)

        self.create_account_button = tk.Button(master, text="Create Account", command=self.open_create_account_window)
        self.create_account_button.pack(pady=5)

        self.reset_password_button = tk.Button(master, text="Reset Password", command=self.reset_password)
        self.reset_password_button.pack(pady=5)  # Add this line for the reset password button

    def ensure_accounts_file_exists(self):
        if not os.path.exists(self.accounts_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Username", "Password", "Security Code"])  # Ensure Security Code column exists
            wb.save(self.accounts_file)
        else:
            # Attempt to open the existing file to check for corruption
            try:
                wb = openpyxl.load_workbook(self.accounts_file)
                ws = wb.active
                # Check if the expected headers are present
                headers = [cell.value for cell in ws[1]]
                if headers != ["Username", "Password", "Security Code"]:
                    raise ValueError("Corrupted file: headers do not match expected format.")
            except Exception as e:
                messagebox.showerror("Error", f"Accounts file is corrupted: {e}. A new file will be created.")
                # Create a new file if the existing one is corrupted
                wb = Workbook()
                ws = wb.active
                ws.append(["Username", "Password", "Security Code"])  # Ensure Security Code column exists
                wb.save(self.accounts_file)

    def account_exists(self, username):
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                return True
        return False

    def add_account(self, username, password, security_code):  # Updated method signature
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        ws.append([username, password, security_code])  # Save security code in the same row
        wb.save(self.accounts_file)  # Ensure the workbook is saved after adding the account

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        if self.check_credentials(username, password):
            self.master.destroy()  # Close login window
            self.open_money_management(username)
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")

    def check_credentials(self, username, password):
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                return True
        return False

    def open_create_account_window(self):
        create_account_window = tk.Toplevel(self.master)
        CreateAccountWindow(create_account_window, self)

    def open_money_management(self, username):
        root = tk.Tk()
        app = MoneyManagementApp(root, username, self.accounts_file)  # Pass accounts_file here
        root.mainloop()

    def reset_password(self):
        username = simpledialog.askstring("Reset Password", "Enter your username:")
        if username and self.account_exists(username):
            attempts = 0
            lockout_time = None
            
            while True:  # Loop until the user successfully resets the password or is locked out
                if lockout_time and time.time() < lockout_time:
                    # User is still locked out
                    remaining_time = int(lockout_time - time.time())
                    messagebox.showwarning("Locked Out", f"You are locked out. Please try again in {remaining_time} seconds.")
                    time.sleep(remaining_time)  # Wait until lockout period is over
                    lockout_time = None  # Reset lockout time after waiting

                security_code = simpledialog.askstring("Security Code", "Enter your 4-digit security code:")
                if self.verify_security_code(username, security_code):
                    new_password = simpledialog.askstring("Reset Password", "Enter your new password:", show="*")
                    if new_password:
                        self.update_password(username, new_password)
                        messagebox.showinfo("Success", "Password has been reset successfully.")
                    return  # Exit the method after successful password reset
                else:
                    attempts += 1
                    messagebox.showerror("Error", f"Invalid security code. You have {3 - attempts} attempts left.")
                    
                    if attempts >= 3:
                        # Lockout logic
                        lockout_time = time.time() + 600  # 10 minutes lockout
                        messagebox.showwarning("Locked Out", "Too many failed attempts. You are locked out for 10 minutes.")
                        break  # Exit the loop after locking out the user
        else:
            messagebox.showerror("Error", "Username does not exist.")

    def verify_security_code(self, username, security_code):
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == username:  # Check if the username matches
                if len(row) > 2 and row[2].value == security_code:  # Check if the security code exists and matches
                    return True
        return False

    def update_password(self, username, new_password):
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == username:
                row[1].value = new_password  # Update the password
                break
        wb.save(self.accounts_file)  # Save changes to the file

    def save_security_code(self, security_code):
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == self.username:  # Match the username
                if len(row) > 2:  # Check if the row has at least 3 columns
                    row[2].value = security_code  # Update the security code in the third column
                else:
                    messagebox.showerror("Error", "User data is corrupted. Please check the accounts file.")
                break
        else:
            messagebox.showerror("Error", "Username not found.")
        wb.save(self.accounts_file)  # Save changes to the file

class MoneyManagementApp:
    def __init__(self, master, username, accounts_file):  # Add accounts_file as a parameter
        self.master = master
        self.master.title("Money Management")
        self.master.geometry("600x700")
        
        self.username = username
        self.accounts_file = accounts_file  # Set the accounts_file attribute
        self.balance = 0
        self.transactions = []
        self.income = {}
        self.expenses = {}
        self.chart_window = None

        # Add username display at the top
        self.username_label = tk.Label(master, text=f"Welcome, {self.username}!", font=("Arial", 16, "bold"))
        self.username_label.pack(pady=10)

        # Create a frame for the reset button
        self.reset_frame = tk.Frame(master)
        self.reset_frame.pack(fill=tk.X, padx=12, pady=6)

        # Add reset button at bottom left
        self.reset_button = tk.Button(self.reset_frame, text="Reset", command=self.reset_data)
        self.reset_button.pack(side=tk.RIGHT)  # Changed from side=tk.RIGHT to side=tk.LEFT
        
        # Create and place widgets
        self.balance_label = tk.Label(master, text=f"Current Balance: ${self.balance}", font=("Arial", 14, "bold"))
        self.balance_label.pack(pady=10)

        self.amount_entry = tk.Entry(master, font=("Arial", 12))
        self.amount_entry.pack()

        self.add_income_button = tk.Button(master, text="Add Income", command=self.add_income)
        self.add_income_button.pack(pady=5)

        self.add_expense_button = tk.Button(master, text="Add Expense", command=self.add_expense)
        self.add_expense_button.pack(pady=5)

        self.transaction_text = tk.Text(master, height=10, width=40, font=("Arial", 10))
        self.transaction_text.pack(pady=10)

        self.show_chart_button = tk.Button(master, text="Show Financial Chart", command=self.show_pie_chart)
        self.show_chart_button.pack(pady=5)

        self.change_bg_button = tk.Button(master, text="Change Background Color", command=self.change_background_color)
        self.change_bg_button.pack(pady=5)

        # Add Instagram link
        self.instagram_button = tk.Button(master, text="Follow us on Instagram", command=self.open_instagram)
        self.instagram_button.pack(pady=10)  # Adjust padding as needed

        # Add Logout button
        self.logout_button = tk.Button(master, text="Logout", command=self.logout)
        self.logout_button.pack(pady=10)  # Adjust padding as needed

        self.username_label = tk.Label(master, text=f"copyright to khellon patel™® © ",font=("Arial", 8))
        self.username_label.pack(pady=50)

        # Add Security Code button
        self.security_code_button = tk.Button(master, text="Set Security Code", command=self.set_security_code)
        self.security_code_button.pack(pady=10)

    def add_income(self):
        self.add_transaction(True)

    def add_expense(self):
        self.add_transaction(False)

    def add_transaction(self, is_income):
        try:
            amount_str = self.amount_entry.get()
            if not amount_str.isdigit():
                raise ValueError("Amount must be a positive integer")
            
            amount = int(amount_str)
            if amount <= 0:
                raise ValueError("Amount must be positive")
            
            if is_income:
                self.balance += amount
                transaction_type = "Income"
                category = simpledialog.askstring("Category", "Enter income category:")
                if category:
                    self.income[category] = self.income.get(category, 0) + amount
            else:
                self.balance -= amount
                transaction_type = "Expense"  # Fixed line
                category = simpledialog.askstring("Category", "Enter expense category:")
                if category:
                    self.expenses[category] = self.expenses.get(category, 0) + amount
            
            self.transactions.append(f"{transaction_type}: ${amount}")
            self.update_display()
            self.amount_entry.delete(0, tk.END)
        except ValueError as e:
            messagebox.showerror("Error", str(e))

    def update_display(self):
        self.balance_label.config(text=f"Current Balance: ${self.balance}")
        self.transaction_text.delete(1.0, tk.END)
        for transaction in self.transactions[-5:]:  # Show last 5 transactions
            self.transaction_text.insert(tk.END, transaction + "\n")

    def show_pie_chart(self):
        if not self.income:
            messagebox.showinfo("Info", "No income data to show")
            return

        # Create a new window for the chart
        chart_window = tk.Toplevel(self.master)
        chart_window.title("Financial Overview")
        chart_window.geometry("600x700")

        # Calculate total income and expenses
        total_income = sum(self.income.values())
        total_expenses = sum(self.expenses.values())

        # Prepare data for the pie chart
        if total_expenses <= total_income:
            sizes = [total_expenses, total_income - total_expenses]
            labels = ['Expenses', 'Remaining Income']
            colors = ['#ff9999', '#66b3ff']  # Light red for expenses, light blue for remaining income
        else:
            sizes = [total_income, total_expenses - total_income]
            labels = ['Income', 'Excess Expenses']
            colors = ['#66b3ff', '#ff9999']  # Light blue for income, light red for excess expenses

        # Create the pie chart
        fig, ax = plt.subplots(figsize=(10, 8))
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors, 
                                          autopct=lambda pct: f"{pct:.1f}%\n(${pct*sum(sizes)/100:.2f})", 
                                          startangle=90, wedgeprops=dict(width=0.5))
        ax.axis('equal')

        # Add a title
        plt.title("Income and Expense Overview", fontsize=16)

        # Add a legend
        ax.legend(wedges, labels,
                  title="Categories",
                  loc="center left",
                  bbox_to_anchor=(1, 0, 0.5, 1))

        # Adjust layout and embed the chart in the Tkinter window
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Add a text widget to display summary and details
        summary_text = tk.Text(chart_window, height=10, width=80)
        summary_text.pack(pady=10)

        # Display summary and details
        summary = f"Total Income: ${total_income}\n"
        summary += f"Total Expenses: ${total_expenses}\n"
        summary += f"Net Balance: ${total_income - total_expenses}\n"
        summary += f"Savings Rate: {(total_income - total_expenses) / total_income * 100:.2f}%\n\n" if total_income > 0 else "Savings Rate: N/A\n\n"
        
        summary += "Income Breakdown:\n"
        for category, amount in self.income.items():
            summary += f"  {category}: ${amount} ({amount/total_income*100:.1f}%)\n"
        
        summary += "\nExpense Breakdown:\n"
        for category, amount in self.expenses.items():
            summary += f"  {category}: ${amount} ({amount/total_income*100:.1f}% of income)\n"

        summary_text.insert(tk.END, summary)
        summary_text.config(state=tk.DISABLED)  # Make the summary read-only

    def change_background_color(self):
        color = colorchooser.askcolor(title="Choose background color")[1]
        if color:
            self.master.configure(bg=color)
            for widget in self.master.winfo_children():
                if isinstance(widget, tk.Label) or isinstance(widget, tk.Text):
                    widget.configure(bg=color)

    def reset_data(self):
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset all data?"):
            self.balance = 0
            self.transactions = []
            self.income = {}
            self.expenses = {}
            self.update_display()
            self.amount_entry.delete(0, tk.END)
            if self.chart_window:
                self.chart_window.destroy()
                self.chart_window = None
            messagebox.showinfo("Reset Complete", "All data has been reset and the chart window has been closed.")

    def open_instagram(self):
        webbrowser.open("https://www.instagram.com/khellon_patel_21")  # Replace with your actual Instagram link

    def logout(self):
        # Logic to return to the login screen
        self.master.destroy()
        main()
        
        
    def set_security_code(self):
        security_code = simpledialog.askstring("Security Code", "Enter a 4-digit security code:")
        if security_code and len(security_code) == 4 and security_code.isdigit():
            self.save_security_code(security_code)
            messagebox.showinfo("Success", "Security code has been set successfully.")
        else:
            messagebox.showerror("Error", "Please enter a valid 4-digit code.")

    def save_security_code(self, security_code):
        wb = openpyxl.load_workbook(self.accounts_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == self.username:  # Match the username
                if len(row) > 2:  # Check if the row has at least 3 columns
                    row[2].value = security_code  # Update the security code in the third column
                else:
                    messagebox.showerror("Error", "User data is corrupted. Please check the accounts file.")
                break
        else:
            messagebox.showerror("Error", "Username not found.")
        wb.save(self.accounts_file)  # Save changes to the file

def main():
    login_root = tk.Tk()
    login_app = LoginWindow(login_root)
    login_root.mainloop()

if __name__ == "__main__":
    main()
